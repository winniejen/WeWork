#!/usr/bin/env python 
# -*- coding: utf-8 -*- 

# @Description:操作excel文件
# @Author : jwy
# @Time : 2021/8/21 19:29
import os

import openpyxl
from openpyxl.styles import Font


class HandleExcel:
    def __init__(self, file_name):
        """
        创建构造函数，初始化类变量
        创建excel工作表对象
        """
        self.file_save_name = file_name.split(".")[0] + "执行结果.xlsx"
        self.font = Font(color=None)  # 设置字体颜色
        self.RGB_dict = {'red': 'FFFF3030', 'green': 'FF008B00'}  # 颜色对应的RGB的值
        # 创建openpyxl工作表
        try:
            self.workbook = openpyxl.load_workbook(file_name)
        except:
            raise Exception

    def get_sheet_by_name(self, sheet_name):
        """
        通过sheet名获取sheet对象
        :param sheet_name: sheet名称
        :return: sheet_name对象
        """
        try:
            sheet = self.workbook[sheet_name]
            return sheet
        except:
            raise Exception

    def get_sheet_by_index(self, sheet_index):
        """
        通过sheet索引号获取sheet对象
        :param sheet_index: sheet索引号
        :return: sheet_index对象
        """
        try:
            sheet = self.workbook[self.workbook[sheet_index]]
            return sheet
        except:
            raise Exception

    def get_row_num(self, sheet):
        """
        获取sheet中有数据区域的结束行号
        :param sheet:
        :return:
        """
        return sheet.max_row

    def get_row_value(self, sheet, row):
        """
        获取sheet中某一行的数据
        :param row: 获取数据的行号
        :param sheet: sheet名称
        :return: 这一行数据内容组成的list
        """
        try:
            row_list = []
            for obj in sheet[row]:
                row_list.append(obj.value)
            return row_list
        except:
            raise Exception

    def get_cell_value(self, sheet, row=None, col=None, coordinate=None):
        """
        获取sheet某一单元格的值
        :param sheet: sheet名称
        :param row: sheet行号
        :param col: sheet列号   # sheet.cell(1,1).value 表示第一行第一列的值
        :param coordinate: 单元格坐标
        :return:
        """
        if coordinate is not None:  # 如果坐标不为空，则根据单元格坐标返回数据
            try:
                return sheet.cell(coordinate=coordinate).value
            except:
                raise Exception
        elif coordinate is None and row is not None and col is not None:  # 如果坐标为空，根据行号和列号返回数据
            try:
                return sheet.cell(row=row, column=col).value
            except:
                raise Exception
        else:
            raise Exception("Insufficient Coordinates of cell!")  # 抛出异常，无效的单元格坐标

    def get_sheet_data(self, sheet):
        """
        获取excel里面所有的数据
        :param sheet: sheet名称
        :return:
        """
        data_list = []
        for i in range(2, self.get_row_num(sheet) + 1):
            data_list.append(self.get_row_value(sheet, i))
        return data_list

    def write_cell(self, sheet, content, coordinate=None, row=None, col=None, style=None):
        """
        写入数据到某一单元格
        :param sheet: sheet名称
        :param content: 写入内容
        :param coordinate: 单元格坐标
        :param row: 单元格行号
        :param col: 单元格列号
        :param style: 写入字体的颜色，red/green
        """
        # 如果坐标不为空，则根据单元格坐标返回数据
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGB_dict[style])
                self.workbook.save(self.file_save_name)
            except:
                raise Exception
        # 如果坐标为空，根据行号和列号写入数据
        elif coordinate is None and row is not None and col is not None:
            try:
                sheet.cell(row=row, column=col).value = content
                if style is not None:
                    sheet.cell(row=row, column=col).font = Font(color=self.RGB_dict[style])
                self.workbook.save(self.file_save_name)
            except:
                raise Exception
        else:
            raise Exception("Insufficient Coordinates of cell!")  # 抛出异常，无效的单元格坐标


# 测试代码
if __name__ == '__main__':
    from util.helper import projet_path  # 导入包

    excel_file = os.path.join(projet_path(), "data/wwork_contacts_testcases.xlsx")
    handle_excel = HandleExcel(excel_file)
    sheet_obj = handle_excel.get_sheet_by_name("成员管理")
    print(sheet_obj)
    print(handle_excel.get_row_num(sheet_obj))
    print(handle_excel.get_row_value(sheet_obj, 1))
    print(handle_excel.get_cell_value(sheet_obj, 1, 1))
